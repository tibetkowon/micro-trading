"""
KIS API Excel 명세서 → Markdown 변환 스크립트

사용법:
    python tools/convert_kis_excel.py <엑셀파일들...> [--output-dir docs/kis-api]

    # 단일 파일
    python tools/convert_kis_excel.py 실시간시세.xlsx

    # 여러 파일 (카테고리별 엑셀)
    python tools/convert_kis_excel.py 실시간시세.xlsx 순위분석.xlsx 시세분석.xlsx

    # 와일드카드
    python tools/convert_kis_excel.py kis-excel/*.xlsx

기능:
    1. 엑셀의 각 API 시트를 개별 마크다운 파일로 변환
    2. 공통 헤더(Request Header)는 _공통헤더.md 로 분리
    3. 긴 코드 목록(5개 초과)은 _공통코드.md 로 분리하고 본문에서는 참조만
    4. 카테고리별로 그룹핑된 인덱스 파일(_index.md) 생성

출력 파일명 규칙:
    {TR_ID}_{API명_간략}.md  (예: HOSTCNT0_실시간체결가.md)
    모든 파일은 flat 구조로 하나의 디렉토리에 저장됨 (카테고리 디렉토리 없음)
"""

import argparse
import glob
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl 패키지가 필요합니다. 설치: pip install openpyxl")
    sys.exit(1)


# ── 공통 헤더 필드: 이 필드들은 _공통헤더.md 로 분리됨 ──────────────────
COMMON_HEADER_FIELDS = {
    "content-type", "authorization", "appkey", "appsecret",
    "personalseckey", "tr_id", "tr_cont", "custtype",
    "seq_no", "mac_address", "phone_number", "ip_addr", "gt_uid",
    # WebSocket 공통
    "approval_key",
}

# 공통코드로 분리할 Description 길이 임계값 (줄 수 기준)
CODE_LIST_THRESHOLD = 5


def cell_value(ws, row, col):
    """셀 값을 문자열로 안전하게 반환"""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return ""
    return str(val).strip()


def parse_meta(ws):
    """시트 상단의 기본정보 파싱 (Row 1~14)"""
    meta = {}

    # Row 1: API 제목 (병합셀일 수 있음)
    meta["title"] = cell_value(ws, 1, 2) or cell_value(ws, 1, 1)

    # Row 2~12: 키-값 쌍
    for row in range(2, 15):
        key = cell_value(ws, row, 1)
        # 값은 B열 이후 병합셀일 수 있으므로 B~G 중 첫 비어있지 않은 값
        val = ""
        for col in range(2, 8):
            v = cell_value(ws, row, col)
            if v:
                val = v
                break
        if key:
            meta[key] = val

    return meta


def find_layout_start(ws):
    """'Layout' 또는 '구분' 헤더 행을 찾아 반환"""
    for row in range(1, ws.max_row + 1):
        a_val = cell_value(ws, row, 1)
        if a_val in ("구분", "Layout"):
            # '구분' 이 헤더행이면 이 행이 컬럼 헤더
            if a_val == "구분":
                return row
            # 'Layout' 이면 다음 행이 컬럼 헤더
            next_val = cell_value(ws, row + 1, 1)
            if next_val == "구분":
                return row + 1
    return None


def find_example_start(ws, layout_start):
    """'Example' 섹션 시작 행을 찾아 반환"""
    for row in range(layout_start, ws.max_row + 1):
        for col in range(1, 8):
            v = cell_value(ws, row, col)
            if v and "Example" in v:
                return row
    return ws.max_row + 1


def parse_layout_rows(ws, start_row, end_row):
    """
    Layout 영역의 필드 행을 파싱하여 섹션별로 그룹화.

    Returns:
        dict: {
            "Request Header": [{"element": ..., "korean": ..., "type": ..., "required": ..., "length": ..., "desc": ...}, ...],
            "Request Body": [...],
            "Response Header": [...],
            "Response Body": [...],
        }
    """
    sections = {}
    current_section = None

    for row in range(start_row + 1, end_row):
        section_label = cell_value(ws, row, 1)
        element = cell_value(ws, row, 2)
        korean = cell_value(ws, row, 3)
        type_ = cell_value(ws, row, 4)
        required = cell_value(ws, row, 5)
        length = cell_value(ws, row, 6)
        desc = cell_value(ws, row, 7)

        # 새 섹션 감지
        if section_label and section_label in (
            "Request Header", "Request Body",
            "Response Header", "Response Body",
        ):
            current_section = section_label
            if current_section not in sections:
                sections[current_section] = []

        # 필드 데이터가 있는 행
        if element and current_section:
            sections[current_section].append({
                "element": element,
                "korean": korean,
                "type": type_,
                "required": required,
                "length": length,
                "desc": desc,
            })

    return sections


def parse_examples(ws, example_start):
    """Example 섹션 파싱"""
    examples = {}
    current_label = None

    for row in range(example_start, ws.max_row + 1):
        label = cell_value(ws, row, 1)
        content = ""
        for col in range(2, 20):  # 넓은 범위에서 값 수집
            v = cell_value(ws, row, col)
            if v:
                content += v + " "
        content = content.strip()

        if label and "Example" in label:
            current_label = label
            if content:
                examples[current_label] = content
        elif current_label and content:
            if current_label in examples:
                examples[current_label] += "\n" + content
            else:
                examples[current_label] = content

    return examples


def is_long_code_list(desc):
    """Description이 긴 코드 목록인지 판별"""
    if not desc:
        return False
    lines = [l.strip() for l in desc.split("\n") if l.strip()]
    # 숫자:설명 또는 코드.설명 패턴이 여러 개인 경우
    code_pattern = re.compile(r"^[\dA-Z][\dA-Z]*[\s]*[:：.·\-]", re.MULTILINE)
    matches = code_pattern.findall(desc)
    return len(matches) > CODE_LIST_THRESHOLD or len(lines) > CODE_LIST_THRESHOLD


def extract_code_name(element, korean):
    """공통코드 참조 키 생성"""
    return f"{element} ({korean})" if korean else element


def is_common_header_field(element):
    """공통 헤더 필드인지 확인"""
    return element.lower() in COMMON_HEADER_FIELDS


def generate_markdown(meta, sections, examples, common_codes_refs):
    """
    파싱된 데이터로 마크다운 문서 생성.

    Args:
        meta: 기본정보 dict
        sections: 섹션별 필드 dict
        examples: 예시 dict
        common_codes_refs: 이 API에서 공통코드로 분리된 필드 set
    """
    lines = []

    # ── 제목 및 기본정보 ─────────────────────────
    title = meta.get("title", meta.get("API 명", "Untitled"))
    lines.append(f"# {title}")
    lines.append("")

    # 기본정보 테이블
    info_keys = [
        ("API 통신방식", "통신방식"),
        ("API 명", "API 명"),
        ("API ID", "API ID"),
        ("실전 TR_ID", "실전 TR_ID"),
        ("모의 TR_ID", "모의 TR_ID"),
        ("HTTP Method", "Method"),
        ("URL 명", "URL"),
    ]

    lines.append("| 항목 | 값 |")
    lines.append("|------|---|")
    for key, label in info_keys:
        val = meta.get(key, "")
        if val:
            lines.append(f"| {label} | `{val}` |")
    lines.append("")

    # 도메인 정보
    real_domain = meta.get("실전 Domain", "")
    mock_domain = meta.get("모의 Domain", "")
    if real_domain or mock_domain:
        lines.append(f"- 실전: `{real_domain}`")
        lines.append(f"- 모의: `{mock_domain}`")
        lines.append("")

    # 개요
    overview = meta.get("개요", "")
    if overview:
        lines.append(f"> {overview}")
        lines.append("")

    # ── Request Header ─────────────────────────
    req_header = sections.get("Request Header", [])
    # 공통 헤더 필드 필터링
    custom_headers = [f for f in req_header if not is_common_header_field(f["element"])]

    if custom_headers:
        lines.append("## Request Header")
        lines.append("")
        lines.append("> 공통 헤더는 [_공통헤더.md](_공통헤더.md) 참조")
        lines.append("")
        lines.append("| Element | 한글명 | Type | Required | Description |")
        lines.append("|---------|--------|------|----------|-------------|")
        for f in custom_headers:
            desc = _abbreviate_desc(f["desc"], f["element"], common_codes_refs)
            lines.append(f"| `{f['element']}` | {f['korean']} | {f['type']} | {f['required']} | {desc} |")
        lines.append("")
    else:
        lines.append("## Request Header")
        lines.append("")
        lines.append("> 공통 헤더만 사용. [_공통헤더.md](_공통헤더.md) 참조")
        lines.append("")

    # ── Request Body / Query ─────────────────────
    req_body = sections.get("Request Body", [])
    if req_body:
        lines.append("## Request Body")
        lines.append("")
        lines.append("| Element | 한글명 | Type | Required | Description |")
        lines.append("|---------|--------|------|----------|-------------|")
        for f in req_body:
            desc = _abbreviate_desc(f["desc"], f["element"], common_codes_refs)
            lines.append(f"| `{f['element']}` | {f['korean']} | {f['type']} | {f['required']} | {desc} |")
        lines.append("")

    # ── Response Header ──────────────────────────
    resp_header = sections.get("Response Header", [])
    if resp_header:
        lines.append("## Response Header")
        lines.append("")
        lines.append("| Element | 한글명 | Type | Required | Description |")
        lines.append("|---------|--------|------|----------|-------------|")
        for f in resp_header:
            desc = _abbreviate_desc(f["desc"], f["element"], common_codes_refs)
            lines.append(f"| `{f['element']}` | {f['korean']} | {f['type']} | {f['required']} | {desc} |")
        lines.append("")

    # ── Response Body ────────────────────────────
    resp_body = sections.get("Response Body", [])
    if resp_body:
        lines.append("## Response Body")
        lines.append("")
        lines.append("| Element | 한글명 | Type | Required | Description |")
        lines.append("|---------|--------|------|----------|-------------|")
        for f in resp_body:
            desc = _abbreviate_desc(f["desc"], f["element"], common_codes_refs)
            lines.append(f"| `{f['element']}` | {f['korean']} | {f['type']} | {f['required']} | {desc} |")
        lines.append("")

    # ── Examples ─────────────────────────────────
    if examples:
        lines.append("## Example")
        lines.append("")
        for label, content in examples.items():
            lines.append(f"### {label}")
            lines.append("```json")
            lines.append(content)
            lines.append("```")
            lines.append("")

    return "\n".join(lines)


def _abbreviate_desc(desc, element, common_codes_refs):
    """
    Description을 축약:
    - 긴 코드 목록 → 공통코드 참조로 대체
    - 줄바꿈 → 세미콜론으로 연결
    - 짧은 코드 목록은 유지
    """
    if not desc:
        return ""

    # 공통코드로 분리된 필드인 경우
    if element in common_codes_refs:
        # 짧은 요약 + 참조 링크
        first_line = desc.split("\n")[0].strip()
        if len(first_line) > 50:
            first_line = first_line[:50] + "…"
        return f"{first_line} → [공통코드](_공통코드.md#{element}) 참조"

    # 줄바꿈을 세미콜론으로 변환 (테이블 내 줄바꿈 방지)
    desc = desc.replace("\n", "; ").replace("\r", "")
    # 연속 세미콜론/공백 정리
    desc = re.sub(r";\s*;", ";", desc)
    desc = re.sub(r"\s+", " ", desc)
    # 파이프 문자 이스케이프 (마크다운 테이블 깨짐 방지)
    desc = desc.replace("|", "\\|")

    return desc.strip()


def sanitize_filename(name):
    """파일명에 사용할 수 없는 문자 제거"""
    # 괄호, 공백 등을 언더스코어로
    name = re.sub(r"[\s/\\:*?\"<>|()（）\[\]]+", "_", name)
    # 연속 언더스코어 정리
    name = re.sub(r"_+", "_", name)
    # 앞뒤 언더스코어 제거
    return name.strip("_")


def generate_common_header_md(all_headers):
    """모든 API에서 수집된 공통 헤더로 _공통헤더.md 생성"""
    lines = [
        "# KIS API 공통 헤더",
        "",
        "> 모든 API에서 공통으로 사용되는 Request Header 필드입니다.",
        "> 개별 API 문서에서는 이 파일을 참조합니다.",
        "",
    ]

    # REST 공통 헤더
    rest_headers = {f["element"]: f for f in all_headers if f["element"].lower() != "approval_key"}
    ws_headers = {f["element"]: f for f in all_headers if f["element"].lower() == "approval_key"}

    if rest_headers:
        lines.append("## REST API 공통 헤더")
        lines.append("")
        lines.append("| Element | 한글명 | Type | Required | Description |")
        lines.append("|---------|--------|------|----------|-------------|")
        for elem, f in rest_headers.items():
            desc = f["desc"].replace("\n", "; ").replace("|", "\\|")
            lines.append(f"| `{f['element']}` | {f['korean']} | {f['type']} | {f['required']} | {desc} |")
        lines.append("")

    if ws_headers:
        lines.append("## WebSocket 공통 헤더")
        lines.append("")
        lines.append("| Element | 한글명 | Type | Required | Description |")
        lines.append("|---------|--------|------|----------|-------------|")
        for elem, f in ws_headers.items():
            desc = f["desc"].replace("\n", "; ").replace("|", "\\|")
            lines.append(f"| `{f['element']}` | {f['korean']} | {f['type']} | {f['required']} | {desc} |")
        lines.append("")

    return "\n".join(lines)


def generate_common_codes_md(all_codes):
    """
    수집된 공통코드를 _공통코드.md 로 생성.

    Args:
        all_codes: dict {element_name: {"korean": ..., "desc": ..., "sources": [API명, ...]}}
    """
    if not all_codes:
        return None

    lines = [
        "# KIS API 공통코드",
        "",
        "> 여러 API 응답에서 반복적으로 사용되는 코드 목록입니다.",
        "> 개별 API 문서에서 '공통코드 참조'로 표기된 필드의 상세 값은 이 문서를 확인하세요.",
        "",
        "---",
        "",
    ]

    for element, data in sorted(all_codes.items()):
        lines.append(f"## {element}")
        lines.append("")
        lines.append(f"**한글명:** {data['korean']}")
        lines.append("")
        lines.append(f"**사용 API:** {', '.join(data['sources'])}")
        lines.append("")

        # Description을 코드 테이블로 변환
        desc_lines = data["desc"].split("\n")
        # 코드:설명 패턴 파싱 시도
        code_entries = []
        for line in desc_lines:
            line = line.strip()
            if not line:
                continue
            # 다양한 구분자 패턴: "1:설명", "01.설명", "A : 설명"
            match = re.match(r"^([A-Z0-9_]+(?:\.[A-Z0-9_]*)?)\s*[:：.·\-]\s*(.+)", line, re.IGNORECASE)
            if match:
                code_entries.append((match.group(1), match.group(2)))
            else:
                code_entries.append(("", line))

        if code_entries:
            lines.append("| 코드 | 설명 |")
            lines.append("|------|------|")
            for code, desc in code_entries:
                desc = desc.replace("|", "\\|")
                lines.append(f"| `{code}` | {desc} |")
        else:
            lines.append(data["desc"])
        lines.append("")
        lines.append("---")
        lines.append("")

    return "\n".join(lines)


def generate_index_md(categorized_apis):
    """
    카테고리별로 그룹핑된 _index.md 생성.

    Args:
        categorized_apis: dict {카테고리명: [{tr_id, api_name, api_type, method, filename}, ...]}
    """
    lines = [
        "# KIS API 인덱스",
        "",
        "> 전체 API 목록과 개별 명세서 링크입니다.",
        "> 카테고리는 원본 엑셀 파일명 기준입니다.",
        "",
    ]

    total = sum(len(apis) for apis in categorized_apis.values())
    lines.append(f"**총 {total}개 API** ({len(categorized_apis)}개 카테고리)")
    lines.append("")

    # 목차
    lines.append("## 목차")
    lines.append("")
    for category in categorized_apis:
        anchor = category.replace(" ", "-")
        count = len(categorized_apis[category])
        lines.append(f"- [{category}](#{anchor}) ({count}개)")
    lines.append("")
    lines.append("---")
    lines.append("")

    # 카테고리별 테이블
    for category, apis in categorized_apis.items():
        lines.append(f"## {category}")
        lines.append("")
        lines.append("| # | 통신방식 | API 명 | TR_ID | Method | 명세서 |")
        lines.append("|---|---------|--------|-------|--------|--------|")
        for i, api in enumerate(apis, 1):
            link = f"[명세서]({api['filename']})" if api.get("filename") else "미생성"
            lines.append(f"| {i} | {api.get('api_type', '')} | {api['api_name']} | `{api['tr_id']}` | {api.get('method', '')} | {link} |")
        lines.append("")

    return "\n".join(lines)


def process_workbook(excel_path, output_path, all_common_headers, all_common_codes, categorized_apis, stats):
    """
    단일 엑셀 워크북을 처리하여 마크다운 파일들 생성.
    공통 데이터(헤더, 코드, 인덱스)는 호출자에서 누적 관리.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # 카테고리명 = 엑셀 파일명 (확장자 제외)
    category = Path(excel_path).stem

    # API 목록 시트와 개별 API 시트 분리
    api_sheets = []
    for name in wb.sheetnames:
        if "API" in name and "목록" in name:
            # 목록 시트에서 메타 정보 수집 (통신방식, Method 등)
            pass
        else:
            api_sheets.append(name)

    print(f"\n📋 [{category}] {excel_path}")
    print(f"   발견된 API 시트: {len(api_sheets)}개")

    category_apis = []

    for sheet_name in api_sheets:
        stats["total"] += 1
        ws = wb[sheet_name]

        try:
            # 1. 메타 정보 파싱
            meta = parse_meta(ws)
            tr_id = meta.get("실전 TR_ID", "")
            api_name = meta.get("API 명", sheet_name)

            if not tr_id:
                print(f"  ⚠️  [{sheet_name}] TR_ID를 찾을 수 없어 시트명 사용")
                tr_id = sanitize_filename(sheet_name)

            # 2. Layout 영역 찾기
            layout_start = find_layout_start(ws)
            if layout_start is None:
                print(f"  ⚠️  [{sheet_name}] Layout 섹션을 찾을 수 없음 → 건너뜀")
                stats["skipped"] += 1
                continue

            # 3. Example 영역 찾기
            example_start = find_example_start(ws, layout_start)

            # 4. 필드 파싱
            sections = parse_layout_rows(ws, layout_start, example_start)

            # 5. 공통 헤더 수집
            req_headers = sections.get("Request Header", [])
            for f in req_headers:
                if is_common_header_field(f["element"]):
                    all_common_headers[f["element"]] = f

            # 6. 긴 코드 목록 → 공통코드로 분리
            common_codes_refs = set()
            for section_name, fields in sections.items():
                for f in fields:
                    if is_long_code_list(f["desc"]):
                        elem = f["element"]
                        common_codes_refs.add(elem)
                        if elem not in all_common_codes:
                            all_common_codes[elem] = {
                                "korean": f["korean"],
                                "desc": f["desc"],
                                "sources": [],
                            }
                        all_common_codes[elem]["sources"].append(api_name)

            # 7. Example 파싱
            examples = parse_examples(ws, example_start)

            # 8. 마크다운 생성
            md_content = generate_markdown(meta, sections, examples, common_codes_refs)

            # 9. 파일 저장
            safe_tr_id = sanitize_filename(tr_id)
            short_name = sanitize_filename(api_name)
            filename = f"{safe_tr_id}_{short_name}.md"
            filepath = output_path / filename
            filepath.write_text(md_content, encoding="utf-8")

            # 인덱스용 정보 수집
            category_apis.append({
                "tr_id": tr_id,
                "api_name": api_name,
                "api_type": meta.get("API 통신방식", ""),
                "method": meta.get("HTTP Method", ""),
                "filename": filename,
            })

            stats["success"] += 1
            print(f"  ✅ [{tr_id}] {api_name} → {filename}")

        except Exception as e:
            stats["errors"].append((sheet_name, str(e)))
            print(f"  ❌ [{sheet_name}] 오류: {e}")

    if category_apis:
        categorized_apis[category] = category_apis


def process_all(excel_paths, output_dir):
    """
    여러 엑셀 파일을 처리하여 하나의 출력 디렉토리에 통합.
    """
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    # 전체 누적 데이터
    all_common_headers = {}
    all_common_codes = {}
    categorized_apis = {}  # {카테고리명: [api_info, ...]}
    stats = {"total": 0, "success": 0, "skipped": 0, "errors": []}

    print(f"📂 출력 디렉토리: {output_dir}")
    print(f"📚 처리할 엑셀 파일: {len(excel_paths)}개")
    print("═" * 50)

    for excel_path in excel_paths:
        process_workbook(
            excel_path, output_path,
            all_common_headers, all_common_codes,
            categorized_apis, stats,
        )

    # ── 공통 파일 생성 (모든 엑셀 처리 후 1회) ────────
    print("\n" + "─" * 50)

    if all_common_headers:
        header_md = generate_common_header_md(list(all_common_headers.values()))
        (output_path / "_공통헤더.md").write_text(header_md, encoding="utf-8")
        print(f"📎 _공통헤더.md 생성 ({len(all_common_headers)}개 필드)")

    if all_common_codes:
        codes_md = generate_common_codes_md(all_common_codes)
        if codes_md:
            (output_path / "_공통코드.md").write_text(codes_md, encoding="utf-8")
            print(f"📎 _공통코드.md 생성 ({len(all_common_codes)}개 코드)")

    if categorized_apis:
        index_md = generate_index_md(categorized_apis)
        (output_path / "_index.md").write_text(index_md, encoding="utf-8")
        print(f"📎 _index.md 생성 ({len(categorized_apis)}개 카테고리)")

    # ── 결과 요약 ──────────────────────────────────
    print("\n" + "═" * 50)
    print(f"📊 변환 결과:")
    print(f"   전체: {stats['total']} | 성공: {stats['success']} | 건너뜀: {stats['skipped']} | 오류: {len(stats['errors'])}")
    if stats["errors"]:
        print(f"\n⚠️  오류 발생 시트:")
        for name, err in stats["errors"]:
            print(f"   - {name}: {err}")

    total_size = sum(f.stat().st_size for f in output_path.glob("*.md"))
    api_count = stats["success"]
    print(f"\n💾 생성된 파일 총 크기: {total_size / 1024:.1f} KB")
    print(f"📉 API 1개 조회 시 평균 토큰: ~{int(total_size / api_count / 4) if api_count else 0} 토큰")


def main():
    parser = argparse.ArgumentParser(
        description="KIS API Excel 명세서 → Markdown 변환",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예시:
    # 단일 파일
    python tools/convert_kis_excel.py 실시간시세.xlsx

    # 여러 파일 (카테고리별)
    python tools/convert_kis_excel.py 실시간시세.xlsx 순위분석.xlsx 주문계좌.xlsx

    # 와일드카드
    python tools/convert_kis_excel.py kis-excel/*.xlsx

    # 출력 디렉토리 지정
    python tools/convert_kis_excel.py kis-excel/*.xlsx -o docs/kis-api
        """,
    )
    parser.add_argument(
        "excel",
        nargs="+",
        help="KIS API 엑셀 명세서 파일 경로 (여러 개 가능, 와일드카드 지원)",
    )
    parser.add_argument(
        "--output-dir", "-o",
        default="docs/kis-api",
        help="마크다운 출력 디렉토리 (기본: docs/kis-api)",
    )

    args = parser.parse_args()

    # 와일드카드 확장 (대괄호 등 특수문자 파일명 대응)
    excel_paths = []
    for pattern in args.excel:
        p = Path(pattern)
        if p.exists():
            # 파일이 직접 존재하면 그대로 사용 ([국내주식] 등 대괄호 파일명 대응)
            excel_paths.append(str(p))
        else:
            # glob 패턴으로 시도
            expanded = glob.glob(pattern)
            if expanded:
                excel_paths.extend(expanded)
            else:
                print(f"⚠️  파일을 찾을 수 없습니다: {pattern}")

    if not excel_paths:
        print("❌ 처리할 엑셀 파일이 없습니다.")
        sys.exit(1)

    # 중복 제거 및 정렬
    excel_paths = sorted(set(excel_paths))

    process_all(excel_paths, args.output_dir)


if __name__ == "__main__":
    main()
